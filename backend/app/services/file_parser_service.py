"""Parser para archivos de pedidos (Excel/CSV y OCR de imágenes).

Detecta automáticamente:
- Formato multi-cliente horizontal (KEEL): múltiples bloques CLIENTE/TRANSPORTE/CODIGO...
  ubicados en columnas adyacentes dentro del mismo sheet.
- Formato simple (un cliente por archivo).
- Imágenes/PDF: extrae texto via Google Vision y aplica heurística.
"""
from __future__ import annotations

import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook

SPREADSHEET_EXT = {".xls", ".xlsx", ".xlsm", ".csv"}
IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".tif", ".pdf"}


@dataclass
class ParsedItem:
    code: str | None
    name: str | None
    quantity_boxes: int
    reference: str | None = None
    fracc: int | None = None


@dataclass
class ParsedOrder:
    client_name: str | None
    transport: str | None
    order_date: date | None = None
    items: list[ParsedItem] = field(default_factory=list)
    source_via_ocr: bool = False
    warnings: list[str] = field(default_factory=list)


class OCRAdapter(ABC):
    @abstractmethod
    async def extract_text(self, content: bytes, mime_type: str) -> str: ...


class GoogleVisionAdapter(OCRAdapter):
    def __init__(self) -> None:
        self._client = None

    def _get_client(self):
        if self._client is None:
            try:
                from google.cloud import vision

                self._client = vision.ImageAnnotatorClient()
            except Exception as exc:
                raise HTTPException(
                    status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                    detail=(
                        "OCR service is not configured. Set GOOGLE_APPLICATION_CREDENTIALS "
                        f"to a valid service account JSON. ({exc})"
                    ),
                ) from exc
        return self._client

    async def extract_text(self, content: bytes, mime_type: str) -> str:
        from google.cloud import vision

        client = self._get_client()
        image = vision.Image(content=content)
        response = client.document_text_detection(image=image)
        if response.error.message:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=f"OCR error: {response.error.message}",
            )
        return response.full_text_annotation.text or ""


class FileParserService:
    def __init__(self, ocr: OCRAdapter | None = None) -> None:
        self._ocr = ocr or GoogleVisionAdapter()

    async def parse(self, file: UploadFile) -> list[ParsedOrder]:
        filename = file.filename or ""
        suffix = Path(filename).suffix.lower()
        content = await file.read()
        await file.seek(0)

        if suffix in IMAGE_EXT:
            text = await self._ocr.extract_text(content, file.content_type or "")
            parsed = self._parse_text(text)
            for p in parsed:
                p.source_via_ocr = True
            return parsed

        if suffix in SPREADSHEET_EXT:
            return self._parse_spreadsheet(content, suffix)

        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Unsupported file type: {suffix or '(none)'}",
        )

    def _parse_spreadsheet(self, content: bytes, suffix: str) -> list[ParsedOrder]:
        if suffix == ".csv":
            import pandas as pd

            df = pd.read_csv(BytesIO(content), header=None, dtype=str, keep_default_na=False)
            return self._parse_grid(df.values.tolist())

        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        out: list[ParsedOrder] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            out.extend(self._parse_grid(rows))
        return out

    def _parse_grid(self, rows) -> list[ParsedOrder]:
        def cell(v) -> str:
            return "" if v is None else str(v).strip()

        grid = [[cell(c) for c in r] for r in rows]
        if not grid:
            return []
        ncols = max(len(r) for r in grid)
        for r in grid:
            r.extend([""] * (ncols - len(r)))

        markers: list[tuple[int, int]] = []
        for ri, row in enumerate(grid):
            for ci, val in enumerate(row):
                upper = val.upper()
                if upper == "CLIENTE" or upper.startswith("CLIENTE:") or upper.startswith("CLIENTE "):
                    markers.append((ri, ci))

        if not markers:
            return []

        orders: list[ParsedOrder] = []
        for r, c in markers:
            parsed = self._parse_client_block(grid, r, c)
            if parsed is not None and (parsed.items or parsed.client_name):
                orders.append(parsed)
        return orders

    def _parse_client_block(self, grid, start_row, start_col) -> ParsedOrder | None:
        max_row = len(grid)
        max_col = len(grid[0]) if grid else 0

        client_name = self._neighbor_value(grid, start_row, start_col)
        transport: str | None = None
        order_date: date | None = None
        header_row: int | None = None
        col_codigo = col_detalle = col_ref = col_fracc = col_cantidad = None

        block_right_limit = min(start_col + 8, max_col)

        for r in range(start_row, min(start_row + 15, max_row)):
            for c in range(start_col, block_right_limit):
                val = grid[r][c].upper()
                if val.startswith("TRANSPORTE"):
                    transport = transport or self._neighbor_value(grid, r, c)
                elif val == "FECHA" or val.startswith("FECHA"):
                    raw = self._neighbor_value(grid, r, c)
                    order_date = self._parse_date(raw)
                elif val == "CODIGO":
                    col_codigo = c
                    header_row = r
                elif val == "DETALLE":
                    col_detalle = c
                elif val.startswith("REF"):
                    col_ref = c
                elif val.startswith("FRAC"):
                    col_fracc = c
                elif val.startswith("CANTIDAD"):
                    col_cantidad = c

        if header_row is None or col_codigo is None:
            return ParsedOrder(
                client_name=client_name,
                transport=transport,
                order_date=order_date,
                warnings=[f"No item header row found for client '{client_name}'"],
            )

        items: list[ParsedItem] = []
        empty_streak = 0
        for r in range(header_row + 1, max_row):
            row = grid[r]
            code = row[col_codigo] if col_codigo < len(row) else ""

            if code.upper() in ("CLIENTE", "CODIGO", "TRANSPORTE"):
                break

            if not code:
                empty_streak += 1
                if empty_streak >= 2:
                    break
                continue
            empty_streak = 0

            qty_str = row[col_cantidad] if col_cantidad is not None and col_cantidad < len(row) else ""
            qty = self._parse_int(qty_str)
            if qty is None or qty <= 0:
                continue

            name = row[col_detalle] if col_detalle is not None and col_detalle < len(row) else None
            ref = row[col_ref] if col_ref is not None and col_ref < len(row) else None
            fracc = self._parse_int(row[col_fracc]) if col_fracc is not None and col_fracc < len(row) else None

            items.append(
                ParsedItem(
                    code=code.strip(),
                    name=name.strip() if name else None,
                    quantity_boxes=qty,
                    reference=ref.strip() if ref else None,
                    fracc=fracc,
                )
            )

        return ParsedOrder(
            client_name=client_name,
            transport=transport,
            order_date=order_date,
            items=items,
        )

    @staticmethod
    def _neighbor_value(grid, r: int, c: int) -> str | None:
        row = grid[r]
        for cc in range(c + 1, min(c + 6, len(row))):
            if row[cc]:
                return row[cc]
        for rr in range(r + 1, min(r + 3, len(grid))):
            if c < len(grid[rr]) and grid[rr][c]:
                return grid[rr][c]
        return None

    @staticmethod
    def _parse_int(val) -> int | None:
        if val is None or val == "":
            return None
        try:
            return int(float(str(val).replace(",", ".").strip()))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_date(val) -> date | None:
        if val is None or val == "":
            return None
        if isinstance(val, date):
            return val
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                from datetime import datetime as dt

                return dt.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_text(self, text: str) -> list[ParsedOrder]:
        """OCR fallback. Best-effort heuristic — verify items in UI before confirming."""
        if not text.strip():
            return []

        lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
        client_name: str | None = None
        transport: str | None = None

        for line in lines:
            upper = line.upper()
            if client_name is None and "CLIENTE" in upper and ":" in line:
                client_name = line.split(":", 1)[1].strip() or None
            elif transport is None and "TRANSPORTE" in upper and ":" in line:
                transport = line.split(":", 1)[1].strip() or None

        items: list[ParsedItem] = []
        warnings: list[str] = []
        pattern = re.compile(r"^([A-Z0-9]{2,8})\s+(.+?)\s+(\d{1,5})\s*$")

        for line in lines:
            match = pattern.match(line)
            if match:
                code, name, qty = match.group(1), match.group(2), int(match.group(3))
                items.append(
                    ParsedItem(code=code, name=name.strip(), quantity_boxes=qty)
                )

        if not items:
            warnings.append(
                "OCR extracted text but could not find any product rows. "
                "Please verify and re-upload, or enter the order manually."
            )

        return [
            ParsedOrder(
                client_name=client_name,
                transport=transport,
                items=items,
                warnings=warnings,
            )
        ]
