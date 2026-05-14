"""Parses STOCK AL DIA 2024.xlsx and yields product tuples for the seed.

Re-usable: imported by initial.py for the actual seed, callable standalone
to verify parsing without writing to the database.
"""
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(__file__).parent / "data" / "STOCK_AL_DIA_2024.xlsx"

SECTIONS_ARTICULOS = {
    "ESCOBILLONES": "ESCOBILLONES",
    "ESCOBAS": "ESCOBAS",
    "CEPILLOS": "CEPILLOS",
    "SECADORES": "SECADORES",
    "VARIOS": "BALDES_Y_ACCESORIOS",
    "CABOS": "BASES_Y_CABOS",
    "ESPONJAS": "ESPONJAS",
    "ANDENES BASE MADERA": "ANDENES",
}

PREFIX = {
    "ESCOBILLONES": "ESC",
    "ESCOBAS": "SCB",
    "CEPILLOS": "CEP",
    "SECADORES": "SEC",
    "BALDES_Y_ACCESORIOS": "VAR",
    "BASES_Y_CABOS": "BAS",
    "ESPONJAS": "ESP",
    "ANDENES": "AND",
    "INSUMOS": "INS",
    "MATERIA_PRIMA": "MP",
    "FIBRA": "FIB",
}


def _clean(x):
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None


def _to_int(x):
    if x is None or isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return int(x)
    try:
        return int(float(str(x).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def parse_workbook() -> list[dict]:
    """Returns a list of product dicts with keys:
    code, name, category, origin, pack_size, qty_boxes, qty_units.
    """
    wb = load_workbook(XLSX_PATH, data_only=True)
    counters: dict[str, int] = {}
    seen: set[str] = set()
    out: list[dict] = []

    def code_for(raw, category: str) -> str:
        c = _clean(raw)
        if c:
            base, i = c, 1
            while c in seen:
                i += 1
                c = f"{base}-{i}"
            seen.add(c)
            return c
        counters[category] = counters.get(category, 0) + 1
        c = f"{PREFIX[category]}-{counters[category]:03d}"
        seen.add(c)
        return c

    def add(code_raw, name: str, category: str, origin: str, pack, qb, qu):
        out.append(
            dict(
                code=code_for(code_raw, category),
                name=name,
                category=category,
                origin=origin,
                pack_size=pack or 1,
                qty_boxes=qb or 0,
                qty_units=qu or 0,
            )
        )

    # ARTICULOS — fábrica propia, secciones por encabezado
    ws = wb["ARTICULOS "]
    current_cat = None
    for row in ws.iter_rows(values_only=True):
        first = _clean(row[0]) if row[0] is not None else None
        if first and first.upper() in SECTIONS_ARTICULOS:
            current_cat = SECTIONS_ARTICULOS[first.upper()]
            continue
        if current_cat is None:
            continue
        name = _clean(row[2])
        if not name:
            continue
        type_pref = _clean(row[1])
        full = f"{type_pref} {name}".strip() if type_pref else name
        add(row[0], full, current_cat, "FABRICACION_PROPIA",
            _to_int(row[3]), _to_int(row[4]), _to_int(row[5]))

    # BETA-CHINO — primer bloque Bettanin (Brasil), segundo bloque China
    ws = wb["BETA-CHINO"]
    in_china = False
    for row in ws.iter_rows(values_only=True):
        first = _clean(row[0])
        name = _clean(row[1])
        if first == "STOCK BETTANIN":
            continue
        if name and "ESCOBILLON" in name.upper() and first:
            try:
                if 100 <= int(float(str(first))) <= 999:
                    in_china = True
            except (ValueError, TypeError):
                pass
        if not name:
            continue
        upper = name.upper()
        if "ESPONJA" in upper:
            cat = "ESPONJAS"
        elif "ESCOBILLON" in upper:
            cat = "ESCOBILLONES"
        elif "CEPILLO" in upper:
            cat = "CEPILLOS"
        else:
            cat = "BALDES_Y_ACCESORIOS"
        origin = "IMPORTADO_CHINA" if in_china else "IMPORTADO_BRASIL"
        add(row[0], name, cat, origin,
            _to_int(row[2]), _to_int(row[3]), _to_int(row[4]))

    # BASES
    ws = wb["BASES"]
    for row in ws.iter_rows(values_only=True):
        first = _clean(row[0])
        if first == "BASES":
            continue
        name = _clean(row[2])
        if not name or "TOTAL" in name.upper():
            continue
        full = name if name.upper().startswith("BASE") else f"BASE {name}"
        add(row[0], full, "BASES_Y_CABOS", "FABRICACION_PROPIA",
            _to_int(row[3]), _to_int(row[4]), _to_int(row[5]))

    # MATERIA PRIMA
    ws = wb["MATERIA PRIMA"]
    for row in ws.iter_rows(values_only=True):
        name = _clean(row[2])
        if not name or "TOTAL" in name.upper():
            continue
        add(None, name, "MATERIA_PRIMA", "FABRICACION_PROPIA",
            _to_int(row[3]), _to_int(row[4]), _to_int(row[5]))

    # INSUMOS
    ws = wb["INSUMOS"]
    for row in ws.iter_rows(values_only=True):
        name = _clean(row[2])
        if not name or "TOTAL" in name.upper():
            continue
        add(row[0], name, "INSUMOS", "FABRICACION_PROPIA",
            _to_int(row[3]), _to_int(row[4]), _to_int(row[5]))

    # FIBRA
    ws = wb["FIBRA"]
    for row in ws.iter_rows(values_only=True):
        name = _clean(row[2])
        if not name or "TOTAL" in name.upper():
            continue
        add(row[0], name, "FIBRA", "FABRICACION_PROPIA",
            _to_int(row[3]), _to_int(row[4]), _to_int(row[5]))

    return out


if __name__ == "__main__":
    rows = parse_workbook()
    print(f"TOTAL: {len(rows)} products")
    print("By category:", dict(Counter(r["category"] for r in rows)))
    print("By origin:", dict(Counter(r["origin"] for r in rows)))
    print()
    print("Sample first 10:")
    for r in rows[:10]:
        print(f"  {r['code']:15s} {r['category']:20s} {r['origin']:20s} pack={r['pack_size']:>4} boxes={r['qty_boxes']:>5}  {r['name']}")
